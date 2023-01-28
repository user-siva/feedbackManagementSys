import socket
import pandas as pd
#import asyncio
serv = socket.socket(socket.AF_INET, socket.SOCK_STREAM)


serv.bind(('0.0.0.0', 8000))
serv.listen(5)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):

    from openpyxl import load_workbook

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:

        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:

            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])

            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow,
                index=False, header=False, **to_excel_kwargs)
    writer.save()


while True:
    conn, addr = serv.accept()
    from_client = ''

    while True:
        data = (conn.recv(40960)).decode()
        if not data:
            break
        df_ = []
        df_.append(eval(data))
        df = pd.DataFrame(df_)
        append_df_to_excel('feedback.xlsx', df=df)
        from_client = data

        print(from_client)

    conn.close()
    print('client disconnected')
# 192.168.43.15
